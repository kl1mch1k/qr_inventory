import os

import openpyxl
import qrcode
from PIL import Image
from flask import render_template, request, abort, send_file, jsonify
from flask_login import login_required, current_user
from openpyxl.styles import Alignment
from werkzeug.utils import redirect
# from data.qr_api import
from data import db_session
from data.history import History
from data.objects import Object
from data.places import Place
from flask_app import app
from forms.object_form import ObjForm
from forms.qr_form import QrForm


@app.route('/')
@login_required
def all_objects():
    if current_user.is_admin():
        return redirect('/admin')
    db_sess = db_session.create_session()
    objects = db_sess.query(Object).filter(Object.responsible_id == current_user.id)
    places = db_sess.query(Place)
    return render_template("user/main_page.html", objects=objects, places=places, title='QR-inventory')


@app.route('/<int:id>')
@login_required
def place_objects(id):
    if current_user.is_admin():
        return redirect('/admin')
    db_sess = db_session.create_session()
    objects = db_sess.query(Object).filter(Object.obj_place == id, Object.responsible_id == current_user.id)
    places = db_sess.query(Place)
    return render_template("user/main_page.html", objects=objects, places=places, title='QR-inventory')

@app.route('/download_qr' , methods=['GET', 'POST'])
@login_required
def user_download_qr():
    form = QrForm()
    form.places.choices = ([(0, 'Все')] + [(place.id, place.text) for place in
                                                                        db_session.create_session().query(
                                                                            Place).all()])
    if form.validate_on_submit():
        wb = openpyxl.Workbook()
        ws = wb.worksheets[0]
        ws.column_dimensions["A"].width = 11
        ws.column_dimensions["C"].width = 11

        db_sess = db_session.create_session()
        places = db_sess.query(Place)
        # objects = db_sess.query(Object).filter(Object.id.in_(request.json['id']))
        if 0 in form.places.data:
            objects = db_sess.query(Object).all()
        else:
            objects = db_sess.query(Object).filter(Object.obj_place.in_(form.places.data)).all()
        count = 2
        for obj in objects:
            qr = qrcode.make(obj.serial_number)
            qr.thumbnail((75, 75), Image.ANTIALIAS)
            qr.save(f'temp/temp{str(count)}.png')
            img = openpyxl.drawing.image.Image(f'temp/temp{str(count)}.png')
            obj_data = {'num': obj.serial_number,
                        'name': obj.name}
            if places.get(obj.obj_place):
                obj_data['place'] = places.get(obj.obj_place).text
            else:
                obj_data['place'] = '___________________'
            string = f'Наим.:{obj_data["name"]}\nМесто:{obj_data["place"]}\nИнв.№:{obj_data["num"]}'
            ws.row_dimensions[count].height = 60
            if not count % 2:
                ws.add_image(img, 'A' + str(count // 2))
                text_cell = ws['B' + str(count // 2)]
                text_cell.value = string
            else:
                ws.add_image(img, 'C' + str(count // 2))
                text_cell = ws['D' + str(count // 2)]
                text_cell.value = string
            text_cell.alignment = Alignment(vertical="center")
            count += 1
        wb.save('temp/out.xlsx')

        # wb_in_bytes = open('temp/out.xlsx', 'rb').read()

        wb.close()
        # os.remove('temp/out.xlsx')

        for i in range(2, count):
            os.remove(f'temp/temp{str(i)}.png')
        return send_file('temp/out.xlsx', as_attachment=True)
    return render_template('user/download_qr.html', title='Загрузка QR', form=form)

    # if current_user.is_authenticated:
    #     file = open('temp/temp.xlsx', 'wb')
    #     file.write(base64.b64decode(requests.get(f'http://127.0.0.1:27016/api/get_xlsx_qr').json()['xlsx'][2:-1]))
    #     return send_file('temp/temp.xlsx', as_attachment=True)
    # else:
    #     return redirect('/login')

# @app.route('/add_object', methods=['GET', 'POST'])
# @login_required
# def add_object():
#     form = ObjForm()
#     form.obj_place.choices = ['Не указано'] + [place.text for place in
#                                                db_session.create_session().query(Place).all()]
#     if form.validate_on_submit():
#         db_sess = db_session.create_session()
#         obj = Object(name=form.name.data,
#                      serial_number=form.serial_number.data)
#         if form.obj_place.data != 'Не указано':
#             obj.obj_place = form.obj_place.data
#         db_sess.add(obj)
#         db_sess.commit()
#         return redirect('/')
#     return render_template('add_object.html', title='Добавление объекта', form=form, editing=False)
#

@app.route('/object/<int:id>', methods=['GET', 'POST'])
@login_required
def edit_object(id):
    form = ObjForm()
    form.obj_place.choices = [(str(i), str(i)) for i in (['Не указано'] +
                                                         [place.text for place in
                                                          db_session.create_session().query(Place).all()])]
    if request.method == "GET":
        db_sess = db_session.create_session()
        obj = db_sess.query(Object).filter(Object.id == id).first()
        if str(obj.responsible_id) != str(current_user.id):
            return jsonify({'error': f"{current_user.email} hasn't got enough rights"})
        if obj:
            if db_sess.query(Place).get(obj.obj_place):
                form = ObjForm(obj_place=db_sess.query(Place).get(obj.obj_place).text)
                form.obj_place.choices = [(str(i), str(i)) for i in (['Не указано'] +
                                                                     [place.text for place in
                                                                      db_session.create_session().query(Place).all()])]
            form.name.data = obj.name
            form.serial_number.data = obj.serial_number

        else:
            abort(404)
    if form.validate_on_submit():
        db_sess = db_session.create_session()
        obj = db_sess.query(Object).filter(Object.id == id).first()
        if obj:
            history = History(old_place_id=obj.obj_place,
                              obj_id=obj.id)
            obj.name = form.name.data
            obj.serial_number = form.serial_number.data
            if form.obj_place.data != 'Не указано':
                obj.obj_place = db_sess.query(Place).filter(Place.text == form.obj_place.data).first().id
                history.new_place_id = obj.obj_place
            db_sess.add(history)
            db_sess.commit()
        else:
            abort(404)
        return redirect('/')
    return render_template('user/add_object.html',
                           title='Редактирование',
                           form=form,
                           obj_id=obj.id,
                           editing=True
                           )


@app.route('/object_delete/<int:id>', methods=['GET', 'POST'])
@login_required
def object_delete(id):
    db_sess = db_session.create_session()
    obj = db_sess.query(Object).filter(Object.id == id).first()
    if obj:
        db_sess.delete(obj)
        db_sess.commit()
    else:
        abort(404)
    return redirect('/')


@app.route('/object_history/<int:id>')
@login_required
def object_history(id):
    db_sess = db_session.create_session()
    obj = db_sess.query(Object).filter(Object.id == id).first()
    if obj:
        histories = reversed(db_sess.query(History).filter(History.obj_id == obj.id).all())
        places = db_sess.query(Place)
        return render_template('object_history.html', obj=obj, places=places, histories=histories,
                               title='Просмотр истории')
    else:
        abort(404)
