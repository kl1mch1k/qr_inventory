import base64
import os

import numpy as np
import openpyxl

import qrcode
from PIL import Image
from flask import jsonify, request
from openpyxl.styles import Alignment

from . import db_session
from .objects import Object

from .api_blueprint import blueprint
from .places import Place


@blueprint.route('/api/get_json_qr', methods=['GET'])
def get_json_qr():
    db_sess = db_session.create_session()
    places = db_sess.query(Place)
    objects = db_sess.query(Object).filter(Object.id.in_(request.json['id']))
    data = {}
    for obj in objects:
        qr = qrcode.make(obj.serial_number)
        qr.thumbnail((75, 75), Image.ANTIALIAS)
        qr = np.array(qr).tolist()
        data[obj.id] = {'qr': qr, 'name': obj.name,
                        'serial_number': obj.serial_number}
    return jsonify({
        'qr_codes': data
    })


@blueprint.route('/api/get_xlsx_qr', methods=['GET'])
def get_xlsx_qr():
    wb = openpyxl.Workbook()
    ws = wb.worksheets[0]
    ws.column_dimensions["A"].width = 11
    ws.column_dimensions["C"].width = 11

    db_sess = db_session.create_session()
    places = db_sess.query(Place)
    # objects = db_sess.query(Object).filter(Object.id.in_(request.json['id']))

    objects = db_sess.query(Object)

    count = 2
    for obj in objects:
        qr = qrcode.make(obj.serial_number)
        qr.thumbnail((75, 75), Image.ANTIALIAS)
        qr.save(f'temp/temp{str(count)}.png')
        img = openpyxl.drawing.image.Image(f'temp/temp{str(count)}.png')
        obj_data = {'num': obj.serial_number,
                    'name': obj.name}
        if places.get(obj.obj_place):
            obj_data['place'] = places.get(obj.obj_place).text
        else:
            obj_data['place'] = '___________________'
        string = f'Наим.:{obj_data["name"]}\nМесто:{obj_data["place"]}\nИнв.№:{obj_data["num"]}'
        ws.row_dimensions[count].height = 60
        if not count % 2:
            ws.add_image(img, 'A' + str(count // 2))
            text_cell = ws['B' + str(count // 2)]
            text_cell.value = string
        else:
            ws.add_image(img, 'C' + str(count // 2))
            text_cell = ws['D' + str(count // 2)]
            text_cell.value = string
        text_cell.alignment = Alignment(vertical="center")
        count += 1
    wb.save('temp/out.xlsx')

    wb_in_bytes = open('temp/out.xlsx', 'rb').read()

    wb.close()
    os.remove('temp/out.xlsx')

    for i in range(2, count):
        os.remove(f'temp/temp{str(i)}.png')

    return jsonify({'xlsx': str(base64.b64encode(wb_in_bytes))})
